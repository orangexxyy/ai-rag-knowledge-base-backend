# app/document_loader.py

import os
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from app.config import (
    OCR_PROVIDER,
    PDF_OCR_MIN_TEXT_CHARS,
    PDF_SCAN_IMAGE_AREA_RATIO,
)
from app.document_models import Document


def format_excel_cell_value(value) -> str:
    """
    将 Excel 单元格值转换成适合拼接进文本的字符串。

    Excel 单元格可能是字符串、数字、日期或空值，统一转换后再进入 RAG。
    """

    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    return str(value).strip()


def build_excel_row_text(
    sheet_name: str,
    headers: list[str],
    row_values: list[str],
) -> str:
    """
    将 Excel 的一行数据转换成自然语言文本。

    Excel 的语义通常来自 header + row，因此不要把整张表简单拼成长文本。
    """

    field_texts = []

    for header, value in zip(headers, row_values):
        if not header or not value:
            continue

        field_texts.append(f"{header}：{value}")

    if not field_texts:
        return ""

    return f"{sheet_name}记录：" + "；".join(field_texts) + "。"


def normalize_pdf_table_cell(value) -> str:
    """
    将 PDF 表格单元格值转成稳定文本。

    pdfplumber 解析表格时可能返回 None、带换行的字符串或数字。
    这里先做轻量规整，避免表格 Document 出现大量空字段或断行噪声。
    """

    if value is None:
        return ""

    return " ".join(str(value).replace("\r", "\n").split()).strip()


def build_pdf_table_row_text(
    page_number: int,
    table_index: int,
    row_index: int,
    headers: list[str],
    row_values: list[str],
) -> str:
    """
    将 PDF 表格中的一行转换成适合 RAG 检索的自然语言 Document 文本。

    设计重点：
    - 不把整张表简单拼成长段文本
    - 保留 page / table / row 语义
    - 用 header=value 让检索词更容易命中具体字段
    """

    field_texts = []

    for column_index, value in enumerate(row_values):
        if not value:
            continue

        if column_index < len(headers) and headers[column_index]:
            header = headers[column_index]
        else:
            header = f"字段{column_index + 1}"

        field_texts.append(f"{header}={value}")

    if not field_texts:
        return ""

    return (
        f"PDF表格记录：第{page_number}页 表格{table_index} "
        f"第{row_index}行："
        + "；".join(field_texts)
        + "。"
    )


def _build_pdf_base_metadata(file_path: str, page_number: int) -> dict:
    """
    统一构造 PDF Document 的基础 metadata，避免 text/table/OCR 分支字段不一致。
    """

    return {
        "source_file": Path(file_path).name,
        "source_path": file_path,
        "file_type": "pdf",
        "page": page_number,
        "sheet_name": None,
        "row_number": None,
        "section_title": None,
        "version": None,
        "permission_level": "internal",
    }


def extract_pdf_table_documents(file_path: str) -> list[Document]:
    """
    使用 pdfplumber 提取文本型 PDF 中可解析的表格，并按 header + row 生成 Document。

    这是最小版表格闭环：
    - 只处理 pdfplumber 能识别出来的文本型表格
    - 不做复杂表格区域去重、跨页表格合并或版面还原
    - 和原始 page text Document 可以少量重复，优先保证结构化表格行可进入 RAG
    """

    try:
        import pdfplumber
    except ImportError:
        print("[WARN] 未安装 pdfplumber，跳过 PDF 表格提取。")
        return []

    table_documents = []

    try:
        with pdfplumber.open(file_path) as pdf:
            for page_index, page in enumerate(pdf.pages):
                page_number = page_index + 1
                tables = page.extract_tables() or []

                for table_index, table in enumerate(tables, start=1):
                    if not table or len(table) < 2:
                        continue

                    headers = [
                        normalize_pdf_table_cell(value) or f"字段{index + 1}"
                        for index, value in enumerate(table[0])
                    ]

                    for row_index, raw_row in enumerate(table[1:], start=1):
                        row_values = [
                            normalize_pdf_table_cell(value)
                            for value in raw_row
                        ]

                        if not any(row_values):
                            continue

                        text = build_pdf_table_row_text(
                            page_number=page_number,
                            table_index=table_index,
                            row_index=row_index,
                            headers=headers,
                            row_values=row_values,
                        )

                        if not text:
                            continue

                        metadata = _build_pdf_base_metadata(
                            file_path=file_path,
                            page_number=page_number,
                        )
                        metadata.update(
                            {
                                "content_type": "table",
                                "table_index": table_index,
                                "row_index": row_index,
                                "row_number": row_index,
                                "extraction_method": "pdfplumber_table",
                            }
                        )

                        table_documents.append(
                            Document(text=text, metadata=metadata)
                        )

    except Exception as exc:
        print(f"[WARN] PDF 表格提取失败，已跳过：{file_path}，原因：{exc}")

    return table_documents


def _get_pypdf_page_image_count(page) -> int:
    """
    从 pypdf PageObject 中尽量读取图片数量。

    该信息只作为最小扫描页启发式信号，不等于生产级扫描页检测。
    """

    try:
        return len(page.images)
    except Exception:
        return 0


def _get_fitz_page_image_stats(
    file_path: str,
    page_index: int,
) -> tuple[int | None, float | None]:
    """
    用 PyMuPDF 估算页面图片数量和图片覆盖面积占比。

    默认 OCR_PROVIDER=none 时不会 import OCR 引擎；这里的 fitz 只用于轻量扫描页判断。
    如果 PyMuPDF 未安装或解析失败，返回 None，由 pypdf 图片数量兜底。
    """

    try:
        import fitz
    except ImportError:
        return None, None

    try:
        with fitz.open(file_path) as pdf:
            page = pdf[page_index]
            page_area = float(page.rect.width * page.rect.height)
            if page_area <= 0:
                return None, None

            image_refs = page.get_images(full=True)
            image_area = 0.0

            for image_ref in image_refs:
                xref = image_ref[0]
                for rect in page.get_image_rects(xref):
                    image_area += float(rect.width * rect.height)

            return len(image_refs), min(image_area / page_area, 1.0)

    except Exception:
        return None, None


def should_handle_pdf_page_image(
    text: str,
    image_count: int,
    image_area_ratio: float | None = None,
) -> bool:
    """
    判断页面是否应进入 OCR / image_placeholder 最小闭环。

    为避免 logo、图标、装饰图片污染索引，必须同时满足：
    - 普通文本很少
    - 页面图片覆盖面积较大，或在无法估算面积时完全没有文本且存在图片
    """

    text_length = len((text or "").strip())

    if text_length > PDF_OCR_MIN_TEXT_CHARS:
        return False

    if image_area_ratio is not None:
        return image_count > 0 and image_area_ratio >= PDF_SCAN_IMAGE_AREA_RATIO

    return text_length == 0 and image_count > 0


def _render_pdf_page_to_array(file_path: str, page_index: int):
    """
    使用 PyMuPDF 将 PDF 页面渲染成图片数组，供可选 OCR Provider 使用。
    """

    import fitz
    import numpy as np

    with fitz.open(file_path) as pdf:
        page = pdf[page_index]
        pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        image = np.frombuffer(pixmap.samples, dtype=np.uint8)
        return image.reshape(pixmap.height, pixmap.width, pixmap.n)


def _run_paddleocr(image_array) -> tuple[str, float | None]:
    """
    运行 PaddleOCR。只有 OCR_PROVIDER=paddleocr 时才会调用并动态 import。
    """

    from paddleocr import PaddleOCR

    ocr = PaddleOCR(use_angle_cls=True, lang="ch")
    result = ocr.ocr(image_array, cls=True) or []

    texts = []
    scores = []

    for page_result in result:
        for line in page_result or []:
            try:
                text = line[1][0]
                score = float(line[1][1])
            except Exception:
                continue

            if text:
                texts.append(text)
                scores.append(score)

    confidence = sum(scores) / len(scores) if scores else None
    return "\n".join(texts).strip(), confidence


def _run_easyocr(image_array) -> tuple[str, float | None]:
    """
    运行 EasyOCR。只有 OCR_PROVIDER=easyocr 时才会调用并动态 import。
    """

    import easyocr

    reader = easyocr.Reader(["ch_sim", "en"], gpu=False)
    result = reader.readtext(image_array) or []

    texts = []
    scores = []

    for item in result:
        try:
            text = item[1]
            score = float(item[2])
        except Exception:
            continue

        if text:
            texts.append(text)
            scores.append(score)

    confidence = sum(scores) / len(scores) if scores else None
    return "\n".join(texts).strip(), confidence


def run_pdf_page_ocr(
    file_path: str,
    page_index: int,
    ocr_provider: str,
) -> tuple[str, float | None, str]:
    """
    对单页 PDF 执行可选 OCR，并返回文本、平均置信度和状态。

    默认 provider=none 时不会调用本函数；paddleocr/easyocr 也只在这里动态 import。
    """

    provider = (ocr_provider or "none").lower()

    if provider not in {"paddleocr", "easyocr"}:
        return "", None, "disabled"

    try:
        image_array = _render_pdf_page_to_array(file_path, page_index)

        if provider == "paddleocr":
            text, confidence = _run_paddleocr(image_array)
        else:
            text, confidence = _run_easyocr(image_array)

    except ImportError as exc:
        return "", None, f"provider_not_installed:{exc.name}"
    except Exception as exc:
        return "", None, f"failed:{exc}"

    if not text.strip():
        return "", confidence, "empty"

    return text.strip(), confidence, "success"


def build_pdf_image_placeholder_document(
    file_path: str,
    page_number: int,
    ocr_provider: str,
    ocr_status: str,
    image_count: int,
    image_area_ratio: float | None = None,
) -> Document:
    """
    构造图片占位 Document。

    只有疑似扫描页 / 主要图片页才会调用它；普通 logo、图标不应进入这里。
    """

    metadata = _build_pdf_base_metadata(
        file_path=file_path,
        page_number=page_number,
    )
    metadata.update(
        {
            "content_type": "image_placeholder",
            "extraction_method": "pypdf_image_detection",
            "ocr_provider": ocr_provider,
            "ocr_status": ocr_status,
            "ocr_confidence": None,
            "image_index": 1,
            "image_count": image_count,
            "image_area_ratio": image_area_ratio,
        }
    )

    return Document(
        text=(
            f"PDF图片占位：第{page_number}页检测到图片内容。"
            "当前未启用 OCR 或多模态图片理解，因此图片中的文字或语义不会进入检索依据。"
        ),
        metadata=metadata,
    )


def evaluate_table_extraction_quality(
    table_rows: list[list[str]],
    ocr_confidence: float | None = None,
    required_headers: list[str] | None = None,
) -> dict:
    """
    评估图片表格 / OCR 表格抽取结果是否适合作为可靠知识入库。

    这是入库阶段 extraction quality control，不是 RAG 检索阶段的 low_confidence。
    当前仅做最小质量门控，用于决定是否需要人工复核。
    """

    reasons = []
    score = 1.0
    required_headers = required_headers or []

    if ocr_confidence is not None and ocr_confidence < 0.6:
        reasons.append(f"OCR 平均置信度较低：{ocr_confidence:.2f}")
        score -= 0.25

    if not table_rows:
        reasons.append("未识别到表格行")
        return {
            "status": "needs_human_review",
            "score": 0.0,
            "reasons": reasons,
            "table_structure_score": 0.0,
        }

    normalized_rows = [
        [normalize_pdf_table_cell(cell) for cell in row]
        for row in table_rows
    ]
    headers = normalized_rows[0] if normalized_rows else []
    non_empty_headers = [header for header in headers if header]

    if not non_empty_headers:
        reasons.append("表头缺失")
        score -= 0.25

    row_lengths = [len(row) for row in normalized_rows if row]
    expected_columns = max(row_lengths, key=row_lengths.count) if row_lengths else 0
    inconsistent_rows = [
        index + 1
        for index, row in enumerate(normalized_rows)
        if len(row) != expected_columns
    ]

    if expected_columns <= 1:
        reasons.append("表格列数过少，结构不稳定")
        score -= 0.2
    elif inconsistent_rows:
        reasons.append(f"行列数不一致：第 {inconsistent_rows} 行")
        score -= 0.2

    total_cells = sum(len(row) for row in normalized_rows)
    empty_cells = sum(1 for row in normalized_rows for cell in row if not cell)
    empty_ratio = empty_cells / total_cells if total_cells else 1.0

    if empty_ratio > 0.4:
        reasons.append(f"空单元格比例过高：{empty_ratio:.2f}")
        score -= 0.2

    missing_required_headers = [
        header for header in required_headers if header not in non_empty_headers
    ]

    if missing_required_headers:
        reasons.append(
            "关键字段缺失：" + "、".join(missing_required_headers)
        )
        score -= 0.25

    score = max(0.0, min(1.0, score))
    table_structure_score = max(
        0.0,
        min(1.0, 1.0 - min(empty_ratio, 1.0) - (0.2 if inconsistent_rows else 0.0)),
    )
    status = "high_confidence" if score >= 0.7 and not reasons else "needs_human_review"

    return {
        "status": status,
        "score": round(score, 4),
        "reasons": reasons,
        "table_structure_score": round(table_structure_score, 4),
    }


def build_pdf_image_table_review_required_document(
    file_path: str,
    page_number: int,
    quality_result: dict,
    ocr_confidence: float | None = None,
    extraction_method: str = "image_table_quality_gate",
) -> Document:
    """
    构造图片表格待人工复核 Document。

    低可信图片表格不会作为可靠 table Document 入库，只保留复核提示和原因。
    """

    reasons = quality_result.get("reasons", [])
    review_reason = "；".join(reasons) if reasons else "抽取质量不足，需要人工复核"

    metadata = _build_pdf_base_metadata(
        file_path=file_path,
        page_number=page_number,
    )
    metadata.update(
        {
            "content_type": "image_table_review_required",
            "human_review_required": True,
            "review_reason": review_reason,
            "extraction_quality_score": quality_result.get("score"),
            "ocr_confidence": ocr_confidence,
            "table_structure_score": quality_result.get("table_structure_score"),
            "extraction_method": extraction_method,
        }
    )

    return Document(
        text=(
            f"PDF图片表格待人工复核：第{page_number}页检测到图片表格，"
            "但 OCR / 表格结构识别可信度较低，未作为可靠知识入库。"
            f"原因：{review_reason}"
        ),
        metadata=metadata,
    )


def build_pdf_image_table_documents_with_quality_gate(
    file_path: str,
    page_number: int,
    table_rows: list[list[str]],
    ocr_confidence: float | None = None,
    required_headers: list[str] | None = None,
    extraction_method: str = "image_table_ocr",
) -> list[Document]:
    """
    将图片表格抽取结果通过质量门控转换成 Document。

    当前项目尚未接入真实图片表格识别/VLM；该函数用于沉淀最小质量门控：
    - 高可信结果才可生成 content_type=table
    - 低可信结果只生成 image_table_review_required，避免污染可靠事实库
    """

    quality_result = evaluate_table_extraction_quality(
        table_rows=table_rows,
        ocr_confidence=ocr_confidence,
        required_headers=required_headers,
    )

    if quality_result["status"] == "needs_human_review":
        return [
            build_pdf_image_table_review_required_document(
                file_path=file_path,
                page_number=page_number,
                quality_result=quality_result,
                ocr_confidence=ocr_confidence,
                extraction_method=extraction_method,
            )
        ]

    headers = [normalize_pdf_table_cell(cell) for cell in table_rows[0]]
    documents = []

    for row_index, raw_row in enumerate(table_rows[1:], start=1):
        row_values = [normalize_pdf_table_cell(cell) for cell in raw_row]
        text = build_pdf_table_row_text(
            page_number=page_number,
            table_index=1,
            row_index=row_index,
            headers=headers,
            row_values=row_values,
        )

        if not text:
            continue

        metadata = _build_pdf_base_metadata(
            file_path=file_path,
            page_number=page_number,
        )
        metadata.update(
            {
                "content_type": "table",
                "table_index": 1,
                "row_index": row_index,
                "row_number": row_index,
                "extraction_method": extraction_method,
                "extraction_quality_score": quality_result.get("score"),
                "ocr_confidence": ocr_confidence,
                "human_review_required": False,
            }
        )
        documents.append(Document(text=text, metadata=metadata))

    return documents


def load_excel_document(file_path: str) -> list[Document]:
    """
    读取 xlsx 文件，并按 sheet / row 转换成 Document 列表。

    当前最小版规则：
    - 每个 sheet 的第一行作为表头
    - 从第二行开始，每一行生成一个 Document
    - 空行跳过
    - metadata 保留 sheet_name 和 row_number
    """

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在：{file_path}")

    # data_only=True 表示如果单元格是公式，读取公式计算后的缓存值。
    workbook = load_workbook(file_path, data_only=True)

    documents = []

    for sheet in workbook.worksheets:
        sheet_name = sheet.title

        if sheet.max_row < 2:
            continue

        header_cells = next(
            sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )

        headers = []

        for index, value in enumerate(header_cells):
            header = format_excel_cell_value(value)

            if not header:
                header = f"字段{index + 1}"

            headers.append(header)

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            row_values = [format_excel_cell_value(value) for value in row]

            if not any(row_values):
                continue

            text = build_excel_row_text(
                sheet_name=sheet_name,
                headers=headers,
                row_values=row_values,
            )

            if not text.strip():
                continue

            metadata = {
                "source_file": Path(file_path).name,
                "source_path": file_path,
                "file_type": "xlsx",
                "page": None,
                "sheet_name": sheet_name,
                "row_number": row_number,
                "section_title": None,
                "version": None,
                "permission_level": "internal",
            }

            documents.append(Document(text=text, metadata=metadata))

    if not documents:
        raise ValueError(f"Excel 未读取到有效数据行：{file_path}")

    return documents


def load_documents_from_dir(dir_path: str) -> list[Document]:
    """
    从目录中批量读取企业资料文件，并转换成统一 Document 列表。
    """

    dir_path_obj = Path(dir_path)

    if not dir_path_obj.exists():
        raise FileNotFoundError(f"资料目录不存在：{dir_path}")

    if not dir_path_obj.is_dir():
        raise NotADirectoryError(f"路径不是目录：{dir_path}")

    all_documents = []
    supported_suffixes = {".txt", ".pdf", ".xlsx"}

    for file_path in sorted(dir_path_obj.iterdir(), key=lambda p: p.name.lower()):
        if not file_path.is_file():
            continue

        suffix = file_path.suffix.lower()

        if suffix not in supported_suffixes:
            print(f"[WARN] 跳过不支持的文件：{file_path}")
            continue

        print(f"正在读取资料文件：{file_path}")
        documents = load_document(str(file_path))
        all_documents.extend(documents)

    if not all_documents:
        raise ValueError(f"资料目录中没有读取到任何支持的文档：{dir_path}")

    return all_documents


def load_txt_document(file_path: str) -> list[Document]:
    """
    读取 txt 文件，并转换成统一 Document 结构。
    """

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在：{file_path}")

    with open(file_path, "r", encoding="utf-8") as f:
        text = f.read()

    metadata = {
        "source_file": Path(file_path).name,
        "source_path": file_path,
        "file_type": "txt",
        "page": None,
        "sheet_name": None,
        "row_number": None,
        "section_title": None,
        "version": None,
        "permission_level": "internal",
    }

    return [Document(text=text, metadata=metadata)]


def load_pdf_document(file_path: str) -> list[Document]:
    """
    读取 PDF，并转换成统一 Document 列表。

    当前最小增强版：
    - 保留原有文本型 PDF page text Document
    - 追加 pdfplumber 可解析的表格 row Document
    - 默认 OCR_PROVIDER=none，不执行 OCR
    - 仅在文本很少且疑似扫描/主要图片页时，生成 OCR 文本或 image_placeholder
    """

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在：{file_path}")

    reader = PdfReader(file_path)
    documents = []

    for page_index, page in enumerate(reader.pages):
        page_number = page_index + 1

        text = page.extract_text() or ""

        image_count = _get_pypdf_page_image_count(page)
        fitz_image_count, image_area_ratio = _get_fitz_page_image_stats(
            file_path=file_path,
            page_index=page_index,
        )

        if fitz_image_count is not None:
            image_count = fitz_image_count

        # 原有能力必须保留：文本型 PDF 每页仍生成 page text Document。
        if text.strip():
            metadata = _build_pdf_base_metadata(
                file_path=file_path,
                page_number=page_number,
            )
            metadata.update(
                {
                    "content_type": "text",
                    "extraction_method": "pypdf_text",
                }
            )

            documents.append(Document(text=text, metadata=metadata))

        # 只在疑似扫描页/主要图片页进入 OCR 或占位逻辑，避免 logo 和装饰图片污染索引。
        if not should_handle_pdf_page_image(
            text=text,
            image_count=image_count,
            image_area_ratio=image_area_ratio,
        ):
            continue

        ocr_status = "disabled"

        if OCR_PROVIDER in {"paddleocr", "easyocr"}:
            ocr_text, ocr_confidence, ocr_status = run_pdf_page_ocr(
                file_path=file_path,
                page_index=page_index,
                ocr_provider=OCR_PROVIDER,
            )

            if ocr_text:
                metadata = _build_pdf_base_metadata(
                    file_path=file_path,
                    page_number=page_number,
                )
                metadata.update(
                    {
                        "content_type": "ocr_text",
                        "extraction_method": f"{OCR_PROVIDER}_page_image",
                        "ocr_provider": OCR_PROVIDER,
                        "ocr_status": ocr_status,
                        "ocr_confidence": ocr_confidence,
                    }
                )

                documents.append(
                    Document(
                        text=f"PDF OCR文本：第{page_number}页：{ocr_text}",
                        metadata=metadata,
                    )
                )
                continue

        documents.append(
            build_pdf_image_placeholder_document(
                file_path=file_path,
                page_number=page_number,
                ocr_provider=OCR_PROVIDER,
                ocr_status=ocr_status,
                image_count=image_count,
                image_area_ratio=image_area_ratio,
            )
        )

    # 表格提取放在文本页之后追加，减少对原有 page text Document 顺序的影响。
    documents.extend(extract_pdf_table_documents(file_path))

    if not documents:
        raise ValueError(
            f"PDF 未提取到任何文本、表格或可处理图片页，可能是空文件或当前最小版暂不支持的 PDF：{file_path}"
        )

    return documents


def load_document(file_path: str) -> list[Document]:
    """
    根据文件后缀选择对应 loader。
    """

    suffix = Path(file_path).suffix.lower()

    if suffix == ".txt":
        return load_txt_document(file_path)

    if suffix == ".pdf":
        return load_pdf_document(file_path)

    if suffix == ".xlsx":
        return load_excel_document(file_path)

    raise ValueError(f"暂不支持的文件类型：{suffix}")
